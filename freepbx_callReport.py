## Gets CDR report using GQL, compares it with phone provider call report to find out which user made which call, then it sets the data in a .xlsx file and sends it using smtp.

from gql import gql, Client
from gql.transport.aiohttp import AIOHTTPTransport
import requests, sys
import pandas as pd
from datetime import datetime
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import smtplib
import email
from email.message import EmailMessage
from email.header import decode_header
import datetime
import locale
from dotenv import load_dotenv
import os
import imaplib
from dateutil.relativedelta import relativedelta

# === CONFIGURATION AND VARIABLES 

locale.setlocale(locale.LC_ALL, ("es_ES", "UTF-8"))
x = datetime.datetime.now()
prev_month_date = x - relativedelta(months=1)

mes = prev_month_date.month
mes_nombre = prev_month_date.strftime('%B').capitalize()
mes_str = f"{mes:02d}"

year = str(x.year)

final_report = 'informe_llamadas-' + year + '-' + mes_str + '.xlsx'

# Get Netelip Report

# === LOAD ENV VARIABLES ===
load_dotenv()

EMAIL_ACCOUNT = os.getenv("EMAIL_ACCOUNT")
PASSWORD = os.getenv("EMAIL_PASSWORD")
IMAP_SERVER = os.getenv("IMAP_SERVER")
SUBJECT_FILTER = os.getenv("SUBJECT_FILTER")
#print(mes_str)
OUTPUT_FILENAME = 'informe_llamadas_' + mes_str + '.csv'

# === CONNECT TO IMAP ===
mail = imaplib.IMAP4_SSL(IMAP_SERVER)
mail.login(EMAIL_ACCOUNT, PASSWORD)
mail.select("inbox")

# === SEARCH FOR ALL EMAILS ===
status, messages = mail.search(None, 'ALL')
if status != 'OK' or not messages[0]:
    #print("No messages found.")
    mail.logout()
    exit()

# Get the most recent 50 emails (adjust if needed)
email_ids = messages[0].split()[-50:]

found_attachment = False

for eid in reversed(email_ids):
    status, msg_data = mail.fetch(eid, '(RFC822)')
    if status != 'OK':
        continue

    msg = email.message_from_bytes(msg_data[0][1])
    subject_parts = decode_header(msg["Subject"])
    subject = ''
    for part, encoding in subject_parts:
        if isinstance(part, bytes):
            subject += part.decode(encoding or 'utf-8', errors='ignore')
        else:
            subject += part


    # Check if subject matches
    print("Email subject:", subject)
    if SUBJECT_FILTER in subject:
        for part in msg.walk():
            if part.get_content_maintype() == 'multipart':
                continue
            if part.get("Content-Disposition") is None:
                continue

            filename = part.get_filename()
            if filename:
                decoded_name, enc = decode_header(filename)[0]
                if isinstance(decoded_name, bytes):
                    filename = decoded_name.decode(enc or "utf-8")
                else:
                    filename = decoded_name

                with open(OUTPUT_FILENAME, "wb") as f:
                    f.write(part.get_payload(decode=True))
                ##print(f"Attachment saved as {OUTPUT_FILENAME}")
                found_attachment = True
                break
        if found_attachment:
            break

if not found_attachment:
    print("No matching email with attachment found.")

mail.logout()

load_dotenv()
TOKEN_URI = os.getenv("TOKEN_URI")
API_URI = os.getenv("API_URI")
AUTH_ID = os.getenv("AUTH_ID")
AUTH_SECRET = os.getenv("AUTH_SECRET")
GQL_SCOPE ='gql'
GQL_QUERY_FETCHALL = f'''
{{
  fetchAllCdrs(
    first: 20000
    after: 1
    orderby: duration
    startDate: "2025-{mes_str}-01"
    endDate: "2025-{mes_str}-31"
  ) {{
    cdrs {{
      uniqueid
      calldate
      clid
      cnum
      dst
      duration
      disposition
    }}
    totalCount
    status
  }}
}}
'''

GQL_QUERY_EXTENSIONS=''' 
query {
  fetchAllExtensions(
    first: 20000
  ) {
    extension {
      user {
        name
        extension
      }
    }
    totalCount
    count
  }
} 
'''

#First authenticate
##print('Requesting authentication token...')
token_request_data={'grant_type':'client_credentials','scope':GQL_SCOPE}
r = requests.post(TOKEN_URI, data=token_request_data, auth=(AUTH_ID, AUTH_SECRET))
if 'access_token' not in r.json():
    sys.exit('Failed to get authentication token. Exiting.')

#Now on to GraphQL
##print('Querying PBX for existing extension list...')
reqHeaders = { 'Authorization': 'Bearer ' + r.json()['access_token'] }
transport = AIOHTTPTransport(url=API_URI, headers=reqHeaders)
client = Client(transport=transport, fetch_schema_from_transport=False)

#Queries
result_cdr = client.execute(gql(GQL_QUERY_FETCHALL))
result_extensions = client.execute(gql(GQL_QUERY_EXTENSIONS))

if not result_cdr['fetchAllCdrs']['status'] or result_cdr['fetchAllCdrs']['totalCount'] < 1:
    sys.exit('Failed to get any cdr info, exiting...')
if not result_extensions['fetchAllExtensions']['totalCount'] or result_extensions['fetchAllExtensions']['count'] < 1:
    sys.exit('Failed to get any extension info, exiting...')

extensiones = {
    
    user['user']['extension']: user['user']['name']
    for user in result_extensions['fetchAllExtensions']['extension']
    if user['user']['extension'] is not None and user['user']['name'] is not None

}

#Create dataframe
df_csv = pd.read_csv('informe_llamadas_'+mes_str+'.csv', sep=';', encoding='utf-8')

#Deletes the columns that are not needed
#df_csv = df_csv.drop(['Plan', 'Useragent', 'IP', 'Nº origen'], axis=1)

#Function to remove the '00' that sometimes go before the number 
def clean_dst(phone_number):
    phone_number = str(phone_number).replace(' ', '').replace(',', '').strip()

    if phone_number.startswith('00'):
      phone_number = phone_number[2:]
    return phone_number

#Uses the funcion on the 'Nº llamado column'
#df_csv['Nº llamado'] = df_csv['Nº llamado'].apply(clean_dst)

#Converts strings into dates  
df_csv['Fecha'] = pd.to_datetime(df_csv['Fecha'], errors='coerce', dayfirst=True)
df_csv['Número Llamado'] = df_csv['Número Llamado'].apply(clean_dst)

#Loops in the result of the API call

for llamada_api in result_cdr['fetchAllCdrs']['cdrs']:
    try:
        api_time = pd.to_datetime(llamada_api['calldate'], errors='coerce')
        api_dst = clean_dst(llamada_api['dst'])

        # Buscar coincidencias por destino
        matching_rows = df_csv[df_csv['Número Llamado'] == api_dst]

        # Filtrar dentro de ±60 segundos
        close_matches = matching_rows[
            matching_rows['Fecha'].apply(lambda x: abs((x - api_time).total_seconds()) <= 60)
        ]

        if not close_matches.empty:
            df_csv.loc[close_matches.index, 'Número origen'] = extensiones.get(llamada_api['cnum'], llamada_api['cnum'])

    # Error handle?
    except Exception as e:
        print(f"Error on {llamada_api.get('uniqueid', '')}: {e}")

    if not close_matches.empty:
        extension = llamada_api['cnum']
        nombre = extensiones.get(extension, extension)  # si no se encuentra, dejar el número
        df_csv.loc[close_matches.index, 'cnum'] = nombre

#df_csv.drop(columns=['Usuario'], inplace=True)
# Exports the new data frame into a csv

df_csv['Número Llamado'] = df_csv['Número Llamado'].astype(str)

df_csv.to_excel(final_report, index=False) 

wb = load_workbook(final_report)
ws = wb.active

column_width = 30
alignment = Alignment(horizontal='center')

for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = column_width
    for cell in column_cells:
        cell.alignment = alignment

wb.save(final_report)

##print(df_csv)



# Send by Email

smtp_server = os.getenv("smtp_server")
smtp_port = os.getenv("smtp_port")
smtp_user = os.getenv("smtp_user")
smtp_password = os.getenv("smtp_password")

sender = os.getenv("sender")
recipient = os.getenv("recipient")
subject = 'Reporte de llamadas Netelip del mes de ' + mes_nombre + '.'
body = 'Hola! \nAquí le adjuntamos el informe de llamadas realizadas mediante Netelip del mes de ' + mes_nombre + '.'

msg = EmailMessage()
msg['From'] = sender
msg['To'] = recipient
msg['Subject'] = subject
msg.set_content(body)

with open(final_report, 'rb') as file:
    file_data = file.read()
    file_name = 'informe_llamadas_' + mes_str + '.xlsx'
    msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=file_name)

with smtplib.SMTP_SSL(smtp_server, 465) as server:
    server.login(smtp_user, smtp_password)
    server.send_message(msg)


##print('Email Enviado!')
